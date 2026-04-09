#!/usr/bin/env python3
"""
Advanced Network Monitor Backend
- Device discovery, bandwidth, port scan, ping/latency
- Suspicious activity detection
- Excel/CSV logging
- HTTP API for frontend
"""

import threading, time, socket, subprocess, ipaddress, platform
import json, csv, os, sys
from datetime import datetime
from collections import defaultdict, deque
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import urlparse, parse_qs

# ── auto-install deps ──────────────────────────────────────────────
def ensure(pkg, import_as=None):
    try:
        __import__(import_as or pkg)
    except ImportError:
        os.system(f"pip install {pkg} --break-system-packages -q")

ensure("psutil")
ensure("openpyxl")

import psutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── CONFIG ─────────────────────────────────────────────────────────
PING_TARGETS = [
    ("Gateway",    None),
    ("Google DNS", "8.8.8.8"),
    ("Cloudflare", "1.1.1.1"),
    ("OpenDNS",    "208.67.222.222"),
]
SCAN_PORTS        = [21,22,23,25,53,80,110,143,443,445,1433,3306,3389,5432,6379,8080,8443,27017]
RISKY_PORTS       = {21,23,3389,445,1433,27017}
REFRESH_SEC       = 3
DISCOVERY_INTERVAL= 45
LATENCY_HISTORY   = 60
LOG_DIR           = "network_logs"
API_PORT          = 7788

# ── SUSPICIOUS ACTIVITY RULES ──────────────────────────────────────
RULES = {
    "high_latency":      lambda ms: ms is not None and ms > 500,
    "packet_loss":       lambda loss_pct: loss_pct > 30,
    "risky_port_open":   lambda port: port in RISKY_PORTS,
    "many_open_ports":   lambda count: count >= 5,
    "bandwidth_spike":   lambda mbps: mbps > 100,
    "new_device":        lambda _: True,   # always flag new devices
}

SEVERITY = {
    "high_latency":    "WARN",
    "packet_loss":     "WARN",
    "risky_port_open": "HIGH",
    "many_open_ports": "MEDIUM",
    "bandwidth_spike": "HIGH",
    "new_device":      "INFO",
}

# ── SHARED STATE ───────────────────────────────────────────────────
state = {
    "ping":           {},
    "bandwidth":      {},
    "devices":        [],
    "ports":          {},
    "alerts":         deque(maxlen=200),
    "events_log":     deque(maxlen=1000),
    "last_discovery": 0,
    "known_ips":      set(),
    "lock":           threading.Lock(),
    "stats": {
        "total_alerts": 0,
        "devices_seen": 0,
        "scans_done":   0,
        "uptime_start": time.time(),
    }
}

os.makedirs(LOG_DIR, exist_ok=True)

# ── HELPERS ────────────────────────────────────────────────────────
def get_default_gateway():
    try:
        if platform.system() == "Windows":
            out = subprocess.check_output("ipconfig", text=True)
            for line in out.splitlines():
                if "Default Gateway" in line:
                    p = line.split(":")
                    if len(p) == 2 and p[1].strip():
                        return p[1].strip()
        else:
            out = subprocess.check_output(["ip", "route"], text=True, stderr=subprocess.DEVNULL)
            for line in out.splitlines():
                if line.startswith("default"):
                    return line.split()[2]
    except Exception:
        pass
    return None

def get_local_subnet():
    try:
        gw = get_default_gateway()
        if not gw: return None
        for _, addrs in psutil.net_if_addrs().items():
            for addr in addrs:
                if addr.family == socket.AF_INET:
                    try:
                        net = ipaddress.IPv4Network(f"{addr.address}/{addr.netmask}", strict=False)
                        if ipaddress.ip_address(gw) in net:
                            return str(net)
                    except Exception:
                        pass
    except Exception:
        pass
    return None

def ping_host(ip, timeout=1.5):
    try:
        flag = "-n" if platform.system() == "Windows" else "-c"
        w_flag = ["-w","1000"] if platform.system() == "Windows" else ["-W","1"]
        cmd = ["ping", flag, "1"] + w_flag + [ip]
        t0 = time.perf_counter()
        r  = subprocess.run(cmd, capture_output=True, text=True, timeout=timeout+1)
        ms = round((time.perf_counter()-t0)*1000, 1)
        return ms if r.returncode == 0 else None
    except Exception:
        return None

def scan_ports(ip, ports, timeout=0.4):
    results = {}
    for port in ports:
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
                s.settimeout(timeout)
                results[port] = s.connect_ex((ip, port)) == 0
        except Exception:
            results[port] = False
    return results

def resolve_hostname(ip):
    try:
        return socket.gethostbyaddr(ip)[0]
    except Exception:
        return ""

def now_str():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

# ── ALERT ENGINE ───────────────────────────────────────────────────
def fire_alert(rule, detail, ip=None, severity=None):
    sev = severity or SEVERITY.get(rule, "INFO")
    alert = {
        "ts":       now_str(),
        "rule":     rule,
        "severity": sev,
        "detail":   detail,
        "ip":       ip or "",
    }
    with state["lock"]:
        state["alerts"].appendleft(alert)
        state["events_log"].appendleft(alert)
        state["stats"]["total_alerts"] += 1
    return alert

# ── LOGGING ────────────────────────────────────────────────────────
def log_to_csv(rows, filename):
    path = os.path.join(LOG_DIR, filename)
    write_header = not os.path.exists(path)
    with open(path, "a", newline="") as f:
        if not rows: return
        w = csv.DictWriter(f, fieldnames=rows[0].keys())
        if write_header: w.writeheader()
        w.writerows(rows)

def export_excel(data_type="all"):
    """Generate a formatted Excel workbook with current state data."""
    wb = Workbook()
    wb.remove(wb.active)

    hdr_fill   = PatternFill("solid", start_color="1A1A2E")
    hdr_font   = Font(bold=True, color="00D4FF", name="Arial", size=10)
    title_font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    alt_fill   = PatternFill("solid", start_color="16213E")
    red_fill   = PatternFill("solid", start_color="4A0000")
    warn_fill  = PatternFill("solid", start_color="3D2B00")
    thin = Side(style="thin", color="2A2A4A")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(ws, headers, row=1):
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=row, column=i, value=h)
            c.fill = hdr_fill
            c.font = hdr_font
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border

    def style_row(ws, row_idx, ncols, severity=None):
        fill = red_fill if severity=="HIGH" else (warn_fill if severity=="WARN" else None)
        for col in range(1, ncols+1):
            c = ws.cell(row=row_idx, column=col)
            c.border = border
            c.alignment = Alignment(vertical="center")
            if fill: c.fill = fill

    with state["lock"]:
        alerts_snap  = list(state["alerts"])
        devices_snap = list(state["devices"])
        ping_snap    = dict(state["ping"])
        bw_snap      = dict(state["bandwidth"])
        ports_snap   = dict(state["ports"])

    # ── Sheet 1: Alerts ──
    ws = wb.create_sheet("🚨 Alerts")
    ws.sheet_view.showGridLines = False
    headers = ["Timestamp","Severity","Rule","Detail","IP Address"]
    style_header(ws, headers)
    ws.row_dimensions[1].height = 22
    for ri, a in enumerate(alerts_snap, 2):
        ws.cell(ri,1,a["ts"]); ws.cell(ri,2,a["severity"])
        ws.cell(ri,3,a["rule"]); ws.cell(ri,4,a["detail"]); ws.cell(ri,5,a["ip"])
        style_row(ws, ri, 5, a["severity"])
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 18

    # ── Sheet 2: Devices ──
    ws2 = wb.create_sheet("🖥 Devices")
    ws2.sheet_view.showGridLines = False
    h2 = ["IP Address","Hostname","Status","RTT (ms)","Open Ports","Risky Ports","First Seen"]
    style_header(ws2, h2)
    for ri, d in enumerate(devices_snap, 2):
        ip = d["ip"]
        pmap = ports_snap.get(ip, {})
        open_p  = [str(p) for p,v in pmap.items() if v]
        risky_p = [str(p) for p in open_p if int(p) in RISKY_PORTS]
        ws2.cell(ri,1,ip); ws2.cell(ri,2,d.get("hostname",""))
        ws2.cell(ri,3,"UP"); ws2.cell(ri,4,d.get("rtt",""))
        ws2.cell(ri,5,", ".join(open_p)); ws2.cell(ri,6,", ".join(risky_p))
        ws2.cell(ri,7,now_str())
        sev = "HIGH" if risky_p else None
        style_row(ws2, ri, 7, sev)
    for col,w in zip("ABCDEFG",[18,30,8,10,30,20,22]): ws2.column_dimensions[col].width = w

    # ── Sheet 3: Latency ──
    ws3 = wb.create_sheet("⚡ Latency")
    ws3.sheet_view.showGridLines = False
    h3 = ["Host","IP","Last RTT (ms)","Min (ms)","Max (ms)","Avg (ms)","Loss %","Samples"]
    style_header(ws3, h3)
    for ri, (name, data) in enumerate(ping_snap.items(), 2):
        hist  = [v for v in data.get("history",[]) if v is not None]
        total = len(list(data.get("history",[])))
        loss  = round((total-len(hist))/max(total,1)*100,1)
        last  = data.get("last")
        ws3.cell(ri,1,name); ws3.cell(ri,2,data.get("ip",""))
        ws3.cell(ri,3,last); ws3.cell(ri,4,min(hist) if hist else "")
        ws3.cell(ri,5,max(hist) if hist else ""); ws3.cell(ri,6,round(sum(hist)/len(hist),1) if hist else "")
        ws3.cell(ri,7,loss); ws3.cell(ri,8,total)
        sev = "HIGH" if loss>30 else ("WARN" if loss>10 else None)
        style_row(ws3, ri, 8, sev)
    for col,w in zip("ABCDEFGH",[14,16,14,10,10,10,8,8]): ws3.column_dimensions[col].width = w

    # ── Sheet 4: Bandwidth ──
    ws4 = wb.create_sheet("📶 Bandwidth")
    ws4.sheet_view.showGridLines = False
    h4 = ["Interface","Recv Rate (KB/s)","Send Rate (KB/s)","Total Recv (MB)","Total Sent (MB)","Pkts In","Pkts Out"]
    style_header(ws4, h4)
    for ri,(iface,d) in enumerate(bw_snap.items(),2):
        ws4.cell(ri,1,iface)
        ws4.cell(ri,2,round(d["recv_rate"]/1024,2))
        ws4.cell(ri,3,round(d["sent_rate"]/1024,2))
        ws4.cell(ri,4,round(d["recv_total"]/1048576,2))
        ws4.cell(ri,5,round(d["sent_total"]/1048576,2))
        ws4.cell(ri,6,d["packets_recv"]); ws4.cell(ri,7,d["packets_sent"])
        sev = "HIGH" if d["recv_rate"]>100*1024*1024 or d["sent_rate"]>100*1024*1024 else None
        style_row(ws4, ri, 7, sev)
    for col,w in zip("ABCDEFG",[18,16,16,16,16,12,12]): ws4.column_dimensions[col].width = w

    # ── Sheet 5: Summary ──
    ws5 = wb.create_sheet("📊 Summary", 0)
    ws5.sheet_view.showGridLines = False
    ws5["A1"] = "NETWORK MONITOR REPORT"
    ws5["A1"].font = Font(bold=True, color="00D4FF", name="Arial", size=16)
    ws5["A2"] = f"Generated: {now_str()}"
    ws5["A2"].font = Font(color="888888", name="Arial", size=10)
    summary_data = [
        ("Total Alerts",      state["stats"]["total_alerts"]),
        ("Devices Online",    len(devices_snap)),
        ("Devices Ever Seen", state["stats"]["devices_seen"]),
        ("Port Scans Done",   state["stats"]["scans_done"]),
        ("Uptime (min)",      round((time.time()-state["stats"]["uptime_start"])/60,1)),
        ("Log Generated",     now_str()),
    ]
    for ri,(k,v) in enumerate(summary_data, 4):
        ws5.cell(ri,1,k).font = Font(bold=True, color="CCCCCC", name="Arial")
        ws5.cell(ri,2,v).font = Font(color="00D4FF", name="Arial")
    ws5.column_dimensions["A"].width = 22
    ws5.column_dimensions["B"].width = 24

    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(LOG_DIR, f"network_report_{ts}.xlsx")
    wb.save(path)
    return path

def export_csv_all():
    with state["lock"]:
        alerts  = list(state["alerts"])
        devices = list(state["devices"])
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    a_path = os.path.join(LOG_DIR, f"alerts_{ts}.csv")
    d_path = os.path.join(LOG_DIR, f"devices_{ts}.csv")
    if alerts:
        with open(a_path,"w",newline="") as f:
            w = csv.DictWriter(f, fieldnames=alerts[0].keys()); w.writeheader(); w.writerows(alerts)
    if devices:
        with open(d_path,"w",newline="") as f:
            w = csv.DictWriter(f, fieldnames=["ip","hostname","status","rtt"]); w.writeheader()
            w.writerows([{k:d.get(k,"") for k in ["ip","hostname","status","rtt"]} for d in devices])
    return [a_path, d_path]

# ── BACKGROUND WORKERS ─────────────────────────────────────────────
def ping_worker():
    gw = get_default_gateway()
    for i,(name,ip) in enumerate(PING_TARGETS):
        if ip is None:
            PING_TARGETS[i] = (name, gw or "127.0.0.1")
    while True:
        for name, ip in PING_TARGETS:
            if not ip: continue
            rtt = ping_host(ip)
            with state["lock"]:
                if name not in state["ping"]:
                    state["ping"][name] = {"ip":ip,"history":deque(maxlen=LATENCY_HISTORY),"last":None}
                state["ping"][name]["history"].append(rtt)
                state["ping"][name]["last"] = rtt
            hist  = list(state["ping"][name]["history"])
            total = len(hist)
            valid = [v for v in hist if v is not None]
            loss  = (total-len(valid))/max(total,1)*100
            if rtt is not None and RULES["high_latency"](rtt):
                fire_alert("high_latency", f"{name} ({ip}) RTT={rtt}ms", ip)
            if total >= 5 and RULES["packet_loss"](loss):
                fire_alert("packet_loss", f"{name} ({ip}) loss={loss:.0f}%", ip)
        time.sleep(REFRESH_SEC)

def bandwidth_worker():
    prev      = psutil.net_io_counters(pernic=True)
    prev_time = time.time()
    while True:
        time.sleep(REFRESH_SEC)
        curr      = psutil.net_io_counters(pernic=True)
        curr_time = time.time()
        dt = curr_time - prev_time or 1
        with state["lock"]:
            for iface, c in curr.items():
                p = prev.get(iface)
                if p is None: continue
                sr = max(0,(c.bytes_sent - p.bytes_sent)/dt)
                rr = max(0,(c.bytes_recv - p.bytes_recv)/dt)
                state["bandwidth"][iface] = {
                    "sent_rate":sr,"recv_rate":rr,
                    "sent_total":c.bytes_sent,"recv_total":c.bytes_recv,
                    "packets_sent":c.packets_sent,"packets_recv":c.packets_recv,
                }
                if RULES["bandwidth_spike"](rr/1048576) or RULES["bandwidth_spike"](sr/1048576):
                    fire_alert("bandwidth_spike", f"{iface} recv={rr/1048576:.1f}MB/s sent={sr/1048576:.1f}MB/s", iface)
        prev=curr; prev_time=curr_time

def discovery_worker():
    while True:
        subnet = get_local_subnet()
        if not subnet: time.sleep(DISCOVERY_INTERVAL); continue
        try:
            hosts = list(ipaddress.IPv4Network(subnet,strict=False).hosts())[:254]
            found = []
            def probe(ip_obj):
                ip  = str(ip_obj)
                rtt = ping_host(ip, timeout=0.5)
                if rtt is not None:
                    hn = resolve_hostname(ip)
                    found.append({"ip":ip,"hostname":hn,"status":"up","rtt":round(rtt,1)})
            threads = [threading.Thread(target=probe,args=(h,),daemon=True) for h in hosts]
            for t in threads: t.start()
            for t in threads: t.join(timeout=3)
            found.sort(key=lambda x: ipaddress.ip_address(x["ip"]))

            with state["lock"]:
                known = state["known_ips"]
                for d in found:
                    if d["ip"] not in known:
                        known.add(d["ip"])
                        state["stats"]["devices_seen"] += 1
                        fire_alert("new_device", f"New device: {d['ip']} ({d['hostname'] or 'unknown'})", d["ip"])
                state["devices"]        = found
                state["last_discovery"] = time.time()

            for dev in found:
                ip      = dev["ip"]
                results = scan_ports(ip, SCAN_PORTS, timeout=0.3)
                with state["lock"]:
                    state["ports"][ip] = results
                    state["stats"]["scans_done"] += 1
                open_p  = [p for p,v in results.items() if v]
                risky_p = [p for p in open_p if p in RISKY_PORTS]
                for rp in risky_p:
                    fire_alert("risky_port_open", f"Port {rp} open on {ip}", ip)
                if RULES["many_open_ports"](len(open_p)):
                    fire_alert("many_open_ports", f"{len(open_p)} ports open on {ip}: {open_p}", ip)
        except Exception as e:
            pass
        time.sleep(DISCOVERY_INTERVAL)

# ── HTTP API ───────────────────────────────────────────────────────
class APIHandler(BaseHTTPRequestHandler):
    def log_message(self, *a): pass  # silence

    def send_json(self, data, code=200):
        body = json.dumps(data, default=str).encode()
        self.send_response(code)
        self.send_header("Content-Type","application/json")
        self.send_header("Access-Control-Allow-Origin","*")
        self.send_header("Content-Length", len(body))
        self.end_headers()
        self.wfile.write(body)

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin","*")
        self.send_header("Access-Control-Allow-Methods","GET,POST,OPTIONS")
        self.end_headers()

    def do_GET(self):
        p = urlparse(self.path)
        qs = parse_qs(p.query)

        if p.path == "/api/state":
            with state["lock"]:
                snap = {
                    "ping":    {k:{**v,"history":list(v["history"])} for k,v in state["ping"].items()},
                    "bandwidth":dict(state["bandwidth"]),
                    "devices": list(state["devices"]),
                    "ports":   {ip:{str(port):open_ for port,open_ in pmap.items()} for ip,pmap in state["ports"].items()},
                    "alerts":  list(state["alerts"])[:50],
                    "stats":   dict(state["stats"]),
                    "last_discovery": state["last_discovery"],
                    "gateway": get_default_gateway(),
                    "subnet":  get_local_subnet(),
                }
            self.send_json(snap)

        elif p.path == "/api/export":
            fmt = qs.get("format",["excel"])[0]
            try:
                if fmt == "csv":
                    paths = export_csv_all()
                    self.send_json({"ok":True,"paths":paths,"format":"csv"})
                else:
                    path = export_excel()
                    self.send_json({"ok":True,"path":path,"format":"excel"})
            except Exception as e:
                self.send_json({"ok":False,"error":str(e)},500)

        elif p.path == "/api/logs":
            files = []
            if os.path.isdir(LOG_DIR):
                for fn in sorted(os.listdir(LOG_DIR), reverse=True)[:20]:
                    fp = os.path.join(LOG_DIR, fn)
                    files.append({"name":fn,"size":os.path.getsize(fp),"modified":os.path.getmtime(fp)})
            self.send_json({"files":files})

        elif p.path.startswith("/logs/"):
            fn   = p.path[6:]
            path = os.path.join(LOG_DIR, fn)
            if os.path.isfile(path):
                with open(path,"rb") as f: data = f.read()
                ext = fn.rsplit(".",1)[-1].lower()
                ct  = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if ext=="xlsx" else "text/csv"
                self.send_response(200)
                self.send_header("Content-Type", ct)
                self.send_header("Content-Disposition", f'attachment; filename="{fn}"')
                self.send_header("Access-Control-Allow-Origin","*")
                self.send_header("Content-Length", len(data))
                self.end_headers()
                self.wfile.write(data)
            else:
                self.send_json({"error":"not found"},404)
        else:
            self.send_json({"error":"not found"},404)


def start_api():
    server = HTTPServer(("0.0.0.0", API_PORT), APIHandler)
    server.serve_forever()

# ── ENTRY POINT ────────────────────────────────────────────────────
if __name__ == "__main__":
    print(f"[*] Starting workers...")
    for fn in [ping_worker, bandwidth_worker, discovery_worker]:
        threading.Thread(target=fn, daemon=True, name=fn.__name__).start()
    time.sleep(1)
    print(f"[*] API server on http://localhost:{API_PORT}")
    print(f"[*] Logs dir: {os.path.abspath(LOG_DIR)}")
    print(f"[*] Press Ctrl+C to stop")
    try:
        start_api()
    except KeyboardInterrupt:
        print("\n[*] Stopped.")
